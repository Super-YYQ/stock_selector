from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COLORS = {
    "ink": "173332",
    "header": "164E4A",
    "accent": "E66B4D",
    "gold": "E9B44C",
    "paper": "F5F7F6",
    "soft": "E8F0EE",
    "red": "C83E38",
    "green": "16805C",
    "muted": "657675",
    "white": "FFFFFF",
    "line": "D8E2E0",
}

TOP50_COLUMNS = {
    "rank": "排名",
    "code": "股票代码",
    "name": "股票名称",
    "total_score": "总分",
    "industry": "所属行业",
    "market_board": "市场板块",
    "pct_chg": "今日涨跌幅",
    "return_5d": "近5日涨跌幅",
    "return_20d": "近20日涨跌幅",
    "amount_ratio": "成交额放大倍数",
    "rps20": "RPS20",
    "rps60": "RPS60",
    "sector_score": "板块分",
    "stock_character_score": "股性分",
    "volume_price_score": "量价分",
    "strategy_score": "策略分",
    "risk_penalty": "风险扣分",
    "reason_tags": "入选理由",
    "concepts": "核心概念",
    "limit_up_reason": "涨停线索",
    "risk_tags": "风险提示",
}

TOP10_COLUMNS = {
    "rank": "排名",
    "code": "股票代码",
    "name": "股票名称",
    "total_score": "总分",
    "industry": "所属行业",
    "market_board": "市场板块",
    "pct_chg": "今日涨跌幅",
    "reason_tags": "重点关注理由",
    "concepts": "核心概念",
    "limit_up_reason": "涨停线索",
    "next_day_condition": "次日观察条件",
    "risk_tags": "风险提示",
}

DETAIL_COLUMNS = {
    "rank": "排名",
    "code": "股票代码",
    "name": "股票名称",
    "market_board": "市场板块",
    "sector": "一级行业",
    "industry": "所属行业",
    "concepts": "核心概念",
    "event_tags": "动态题材标签",
    "limit_up_reason": "涨停线索",
    "industry_activity": "行业阶段表现",
    "matched_strategies": "命中策略",
    "strategy_reason": "策略理由",
    "selection_reason": "完整入选理由",
    "stock_context_summary": "行业与题材说明",
    "next_day_condition": "次日观察条件",
    "risk_warning": "完整风险提示",
}

SECTOR_COLUMNS = {
    "sector_name": "板块",
    "sector_score_raw": "板块热度",
    "pct_chg": "今日涨跌幅",
    "pct_chg_5d": "近5日涨跌幅",
    "pct_chg_20d": "近20日涨跌幅",
    "amount_ratio": "成交额放大倍数",
    "limit_up_count": "涨停家数",
    "strong_stock_count": "强势股家数",
    "sector_reason": "强势原因",
}

PERFORMANCE_COLUMNS = {
    "strategy": "策略",
    "sample_count": "样本数",
    "return_1d": "平均1日收益",
    "win_rate_1d": "1日胜率",
    "return_3d": "平均3日收益",
    "win_rate_3d": "3日胜率",
    "return_5d": "平均5日收益",
    "win_rate_5d": "5日胜率",
    "return_10d": "平均10日收益",
    "win_rate_10d": "10日胜率",
}

CUSTOM_FORMULA_COLUMNS = {
    "formula_rank": "公式内排名",
    "custom_strategy_name": "自定义公式",
    "code": "股票代码",
    "name": "股票名称",
    "total_score": "综合分",
    "industry": "所属行业",
    "market_board": "市场板块",
    "pct_chg": "今日涨跌幅",
    "amount_ratio": "成交额放大倍数",
    "rps20": "RPS20",
    "distance_ma20": "距20日线",
    "custom_reason": "公式命中原因",
    "risk_tags": "风险提示",
}

FILTER_COLUMNS = {
    "code": "股票代码",
    "name": "股票名称",
    "industry": "所属板块",
    "filter_reason": "过滤原因",
}


def _rename_existing(df: pd.DataFrame, columns: dict[str, str]) -> pd.DataFrame:
    return df.reindex(columns=list(columns)).rename(columns=columns)


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _column_values(ws, column: int, limit: int = 100) -> Iterable[object]:
    for row in range(1, min(ws.max_row, limit) + 1):
        yield ws.cell(row=row, column=column).value


def _fit_columns(ws) -> None:
    long_headers = {
        "命中策略",
        "入选标签",
        "核心概念",
        "涨停线索",
        "行业阶段表现",
        "行业与题材说明",
        "动态题材标签",
        "策略理由",
        "完整入选理由",
        "完整风险提示",
        "入选理由",
        "重点关注理由",
        "次日观察条件",
        "风险提示",
        "过滤原因",
        "强势原因",
        "公式命中原因",
    }
    for column in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=column).value or "")
        values = [len(str(value)) for value in _column_values(ws, column) if value is not None]
        width = min(max(values or [8]) + 2, 22)
        if header in long_headers:
            width = 34
        elif "代码" in header:
            width = 12
        elif header in {"股票名称", "所属板块", "策略"}:
            width = 16
        ws.column_dimensions[get_column_letter(column)].width = width


def _style_data_sheet(ws) -> None:
    _style_header(ws)
    ws.freeze_panes = "D2" if ws.title in {"Top50观察名单", "Top10重点关注", "个股说明", "自定义策略"} else "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    headers = {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}
    wrap_headers = {
        "入选理由",
        "重点关注理由",
        "核心概念",
        "涨停线索",
        "风险提示",
        "命中策略",
        "动态题材标签",
        "行业阶段表现",
        "策略理由",
        "完整入选理由",
        "行业与题材说明",
        "次日观察条件",
        "完整风险提示",
        "过滤原因",
        "强势原因",
        "公式命中原因",
    }
    wrap_columns = {column for header, column in headers.items() if header in wrap_headers}
    thin = Side(style="hair", color=COLORS["line"])
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = PatternFill("solid", fgColor=COLORS["paper"])
        ws.row_dimensions[row].height = 34 if wrap_columns else 25
        for cell in ws[row]:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left" if cell.column in wrap_columns else "center",
                wrap_text=cell.column in wrap_columns,
            )
    for header, column in headers.items():
        letter = get_column_letter(column)
        if "涨跌幅" in header or "平均" in header or "收益" in header or "胜率" in header:
            for cell in ws[letter][1:]:
                cell.number_format = '0.00"%"'
            if ws.max_row >= 2:
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="greaterThan", formula=["0"], font=Font(color=COLORS["red"])),
                )
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="lessThan", formula=["0"], font=Font(color=COLORS["green"])),
                )
        elif "放大倍数" in header:
            for cell in ws[letter][1:]:
                cell.number_format = '0.00"x"'
        elif header == "总分" and ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F2E7E3",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="F4DFA7",
                    end_type="max",
                    end_color="7DB9A8",
                ),
            )
        elif header in {"RPS20", "RPS60", "板块热度"} and ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{ws.max_row}",
                DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color=COLORS["accent"]),
            )
        elif header == "股票代码":
            for cell in ws[letter][1:]:
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value).zfill(6)
    _fit_columns(ws)


def _write_market_sheet(
    writer: pd.ExcelWriter,
    report_date: str,
    market: dict[str, object],
    health: dict[str, object],
    snapshot_type: str = "close",
) -> None:
    pd.DataFrame().to_excel(writer, sheet_name="市场环境", index=False, header=False)
    ws = writer.sheets["市场环境"]
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    title = ws["A1"]
    title.value = (
        f"{report_date} A股盘中快照"
        if snapshot_type == "intraday"
        else f"{report_date} A股盘后观察"
    )
    title.font = Font(size=22, bold=True, color=COLORS["white"])
    title.fill = PatternFill("solid", fgColor=COLORS["header"])
    title.alignment = Alignment(horizontal="left", vertical="center")
    for row in ws["A1:H2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=COLORS["header"])

    metrics = [
        ("市场环境", market.get("market_label", "-")),
        ("市场评分", market.get("market_score", 0)),
        ("风险等级", market.get("risk_level", "-")),
        ("上涨占比", f"{market.get('up_ratio', 0)}%"),
        ("涨停家数", market.get("limit_up_count", 0)),
        ("跌停家数", market.get("limit_down_count", 0)),
        ("最新交易日", health.get("latest_trade_date", report_date)),
        ("数据覆盖率", f"{float(health.get('stock_coverage', 0)):.1%}"),
    ]
    positions = [(1, 4), (3, 4), (5, 4), (7, 4), (1, 7), (3, 7), (5, 7), (7, 7)]
    for (label, value), (column, row) in zip(metrics, positions):
        ws.cell(row=row, column=column, value=label)
        ws.cell(row=row + 1, column=column, value=value)
        ws.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        ws.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        label_cell = ws.cell(row=row, column=column)
        value_cell = ws.cell(row=row + 1, column=column)
        label_cell.font = Font(color=COLORS["muted"], bold=True)
        value_cell.font = Font(color=COLORS["ink"], bold=True, size=16)
        label_cell.fill = PatternFill("solid", fgColor=COLORS["soft"])
        value_cell.fill = PatternFill("solid", fgColor=COLORS["paper"])
        label_cell.alignment = value_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 23
        ws.row_dimensions[row + 1].height = 31

    index_changes = market.get("index_changes", {}) or {}
    labels = {"sh000001": "上证指数", "sz399001": "深证成指", "sz399006": "创业板指"}
    ws["A11"] = "主要指数"
    ws["A11"].font = Font(size=13, bold=True, color=COLORS["ink"])
    for offset, code in enumerate(labels, start=12):
        ws.cell(offset, 1, labels[code])
        ws.cell(offset, 2, float(index_changes.get(code, 0) or 0))
        ws.cell(offset, 2).number_format = '0.00"%"'
        ws.cell(offset, 2).font = Font(color=COLORS["red"] if float(index_changes.get(code, 0) or 0) >= 0 else COLORS["green"])

    ws.merge_cells("A16:H17")
    ws["A16"] = (
        "本报告为盘中临时快照，当日日K尚未收盘，不写入正式策略收益历史，仅供观察。"
        if snapshot_type == "intraday"
        else "本报告仅用于盘后复盘和次日观察，不构成投资建议，也不执行自动交易。"
    )
    ws["A16"].font = Font(color=COLORS["muted"], italic=True)
    ws["A16"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A16"].fill = PatternFill("solid", fgColor=COLORS["paper"])
    for column in range(1, 9):
        ws.column_dimensions[get_column_letter(column)].width = 15


def write_excel_report(
    output_dir: str | Path,
    report_date: str,
    market: dict[str, object],
    strong_sectors: pd.DataFrame,
    top50: pd.DataFrame,
    top10: pd.DataFrame,
    ranked: pd.DataFrame,
    filtered: pd.DataFrame,
    strategy_performance: pd.DataFrame | None = None,
    health: dict[str, object] | None = None,
    custom_strategy_results: pd.DataFrame | None = None,
    snapshot_type: str = "close",
) -> Path:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    suffix = "盘中选股快照" if snapshot_type == "intraday" else "盘后选股报告"
    path = output / f"{report_date}_{suffix}.xlsx"
    performance = strategy_performance if strategy_performance is not None else pd.DataFrame()
    custom_results = custom_strategy_results if custom_strategy_results is not None else pd.DataFrame()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_market_sheet(writer, report_date, market, health or {}, snapshot_type)
        _rename_existing(strong_sectors, SECTOR_COLUMNS).head(30).to_excel(writer, sheet_name="强势板块", index=False)
        _rename_existing(top50, TOP50_COLUMNS).to_excel(writer, sheet_name="Top50观察名单", index=False)
        _rename_existing(top10, TOP10_COLUMNS).to_excel(writer, sheet_name="Top10重点关注", index=False)
        _rename_existing(top50, DETAIL_COLUMNS).to_excel(writer, sheet_name="个股说明", index=False)
        _rename_existing(performance, PERFORMANCE_COLUMNS).to_excel(writer, sheet_name="策略表现", index=False)
        _rename_existing(custom_results, CUSTOM_FORMULA_COLUMNS).to_excel(writer, sheet_name="自定义策略", index=False)
        _rename_existing(filtered, FILTER_COLUMNS).to_excel(writer, sheet_name="风险过滤名单", index=False)
        ranked.to_excel(writer, sheet_name="原始评分明细", index=False)

        for name in ["强势板块", "Top50观察名单", "Top10重点关注", "个股说明", "策略表现", "自定义策略", "风险过滤名单", "原始评分明细"]:
            _style_data_sheet(writer.sheets[name])
        tab_colors = {
            "市场环境": COLORS["header"],
            "强势板块": COLORS["gold"],
            "Top50观察名单": COLORS["green"],
            "Top10重点关注": COLORS["accent"],
            "个股说明": COLORS["muted"],
            "自定义策略": COLORS["header"],
            "风险过滤名单": COLORS["red"],
        }
        for name, color in tab_colors.items():
            writer.sheets[name].sheet_properties.tabColor = color
        writer.sheets["原始评分明细"].sheet_state = "hidden"
        writer.book.active = writer.book.sheetnames.index("市场环境")
    return path
