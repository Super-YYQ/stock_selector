from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.report import write_excel_report


def test_write_excel_report_creates_required_sheets(tmp_path: Path) -> None:
    market = {
        "market_label": "偏强",
        "risk_level": "中",
        "market_score": 7.5,
        "up_ratio": 62,
        "limit_up_count": 78,
        "limit_down_count": 12,
    }
    strong_sectors = pd.DataFrame([{"sector_name": "机器人", "sector_score_raw": 90, "sector_reason": "板块涨幅 4.2%"}])
    ranked = pd.DataFrame(
        [
            {
                "rank": 1,
                "code": "000001",
                "name": "强势股",
                "industry": "机器人",
                "total_score": 88,
                "pct_chg": 4,
                "return_5d": 8,
                "return_20d": 20,
                "amount_ratio": 2,
                "rps20": 90,
                "rps60": 80,
                "sector_score": 22,
                "stock_character_score": 16,
                "volume_price_score": 21,
                "risk_penalty": 5,
                "selection_reason": "放量突破",
                "next_day_condition": "不追高",
                "risk_warning": "暂无明显量化风险",
            }
        ]
    )
    filtered = pd.DataFrame([{"code": "000002", "name": "ST测试", "filter_reason": "ST 或退市风险"}])

    path = write_excel_report(tmp_path, "2026-06-22", market, strong_sectors, ranked, ranked, ranked, filtered)

    assert path.exists()
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["市场环境", "强势板块", "Top50观察名单", "Top10重点关注", "风险过滤名单", "原始评分明细"]
    assert workbook["Top50观察名单"]["A1"].value == "排名"
